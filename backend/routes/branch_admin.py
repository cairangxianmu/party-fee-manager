import io
import json as _json
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from flask import Blueprint, request, jsonify, g, send_file
from database import get_db
from auth import require_admin, log_action
from .super_admin import _validate_member_payload, _friendly_unique_msg, _PHONE_RE

branch_bp = Blueprint("branch", __name__)


# ============================================================
# 本支部成员管理
# ============================================================

@branch_bp.route("/branch/members", methods=["GET"])
@require_admin
def list_members():
    status = request.args.get("status")
    sql = """SELECT m.*, b.name AS branch_name FROM members m
             LEFT JOIN branches b ON m.branch_id = b.id
             WHERE m.branch_id=?"""
    params = [g.branch_id]
    if status:
        sql += " AND m.status=?"
        params.append(status)
    sql += " ORDER BY m.id"

    db = get_db()
    rows = db.execute(sql, params).fetchall()
    db.close()
    return jsonify({"code": 0, "data": [dict(r) for r in rows]})


@branch_bp.route("/branch/members", methods=["POST"])
@require_admin
def create_member():
    data = request.get_json() or {}
    err = _validate_member_payload(data, require_branch=False)
    if err:
        return jsonify({"code": 400, "msg": err})

    db = get_db()
    try:
        db.execute(
            """INSERT INTO members (name, member_no, phone, identity, person_type, branch_id, amount, status, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                data["name"].strip(), str(data["member_no"]).strip(),
                str(data["phone"]).strip(), data["identity"].strip(),
                (data.get("person_type") or "").strip(),
                g.branch_id, float(data["amount"]),
                data.get("status", "active"), (data.get("notes") or "").strip(),
            ),
        )
        log_action(g.admin_id, "新增成员（支部）", f"姓名：{data['name']}", db=db)
        db.commit()
    except Exception as e:
        db.close()
        return jsonify({"code": 400, "msg": _friendly_unique_msg(e) or "工号/学号或手机号已存在"})
    db.close()
    return jsonify({"code": 0, "msg": "新增成功"})


@branch_bp.route("/branch/members/import", methods=["POST"])
@require_admin
def import_members():
    if "file" not in request.files:
        return jsonify({"code": 400, "msg": "未找到文件"})
    file = request.files["file"]
    file.seek(0, 2)
    if file.tell() > 5 * 1024 * 1024:
        return jsonify({"code": 400, "msg": "文件大小不能超过 5MB"})
    file.seek(0)
    try:
        wb = openpyxl.load_workbook(file)
    except Exception:
        return jsonify({"code": 400, "msg": "文件格式错误，请上传 .xlsx 文件"})

    ws = wb.active
    if ws.max_row - 1 > 1000:
        return jsonify({"code": 400, "msg": "单次导入不能超过 1000 条"})
    db = get_db()
    success_count, skip_count, error_rows = 0, 0, []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        try:
            name        = str(row[0]).strip()
            member_no   = str(row[1]).strip()
            phone       = str(row[2]).strip()
            identity    = str(row[3]).strip()
            person_type = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            # row[5] = 支部名称，支部管理员导入时忽略，强制使用自己的支部
            amount      = float(row[6] or 0) if len(row) > 6 and row[6] is not None else 0
            status_raw  = str(row[7]).strip() if len(row) > 7 and row[7] else "正常"
            notes       = str(row[8]).strip() if len(row) > 8 and row[8] else ""
            _STATUS_MAP = {"正常": "active", "停缴": "suspended", "已转出": "transferred",
                           "active": "active", "suspended": "suspended", "transferred": "transferred"}
            status = _STATUS_MAP.get(status_raw)

            if not name or not member_no or not phone:
                error_rows.append(f"第{i}行：姓名、工号/学号、手机号不能为空")
                continue
            if not _PHONE_RE.match(phone):
                error_rows.append(f"第{i}行：手机号「{phone}」必须为11位数字")
                continue
            if amount <= 0:
                error_rows.append(f"第{i}行：应缴金额须为正数")
                continue
            if amount > 100000:
                error_rows.append(f"第{i}行：应缴金额异常（>100000）")
                continue
            if status is None:
                error_rows.append(f"第{i}行：状态值「{status_raw}」非法，请填写正常/停缴/已转出")
                continue
            amount = round(amount, 2)

            cursor = db.execute(
                """INSERT OR IGNORE INTO members
                   (name, member_no, phone, identity, person_type, branch_id, amount, status, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (name, member_no, phone, identity, person_type, g.branch_id, amount, status, notes),
            )
            if cursor.rowcount > 0:
                success_count += 1
            else:
                skip_count += 1
        except Exception as e:
            error_rows.append(f"第{i}行：{str(e)}")

    log_action(g.admin_id, "批量导入成员（支部）", f"成功{success_count}条，跳过{skip_count}条", db=db)
    db.commit()
    db.close()

    parts = [f"成功导入 {success_count} 条"]
    if skip_count > 0:
        parts.append(f"{skip_count} 条已存在跳过")
    return jsonify({"code": 0, "msg": "，".join(parts), "errors": error_rows})


@branch_bp.route("/branch/members/<int:mid>", methods=["PUT"])
@require_admin
def update_member(mid):
    db = get_db()
    member = db.execute(
        "SELECT * FROM members WHERE id=? AND branch_id=?", (mid, g.branch_id)
    ).fetchone()
    if not member:
        db.close()
        return jsonify({"code": 403, "msg": "无权操作此成员"})

    data = request.get_json() or {}
    # 填充缺省字段后做统一校验（允许前端只传部分字段）
    merged = {
        "name": data.get("name", member["name"]),
        "member_no": data.get("member_no", member["member_no"]),
        "phone": data.get("phone", member["phone"]),
        "identity": data.get("identity", member["identity"]),
        "person_type": data.get("person_type", member["person_type"] or ""),
        "amount": data.get("amount", member["amount"]),
        "status": data.get("status", member["status"]),
        "notes": data.get("notes", member["notes"] or ""),
    }
    err = _validate_member_payload(merged, require_branch=False)
    if err:
        db.close()
        return jsonify({"code": 400, "msg": err})

    try:
        db.execute(
            """UPDATE members SET member_no=?, name=?, phone=?, identity=?, person_type=?,
               amount=?, status=?, notes=? WHERE id=? AND branch_id=?""",
            (
                str(merged["member_no"]).strip(),
                merged["name"].strip(),
                str(merged["phone"]).strip(),
                merged["identity"].strip(),
                (merged["person_type"] or "").strip(),
                float(merged["amount"]),
                merged["status"],
                (merged["notes"] or "").strip(),
                mid, g.branch_id,
            ),
        )
        log_action(g.admin_id, "编辑成员（支部）", f"成员ID：{mid}", db=db)
        db.commit()
    except Exception as e:
        db.close()
        return jsonify({"code": 400, "msg": _friendly_unique_msg(e) or "保存失败：工号/学号或手机号冲突"})
    db.close()
    return jsonify({"code": 0, "msg": "更新成功"})


@branch_bp.route("/branch/members/<int:mid>", methods=["DELETE"])
@require_admin
def delete_member(mid):
    db = get_db()
    member = db.execute(
        "SELECT * FROM members WHERE id=? AND branch_id=?", (mid, g.branch_id)
    ).fetchone()
    if not member:
        db.close()
        return jsonify({"code": 403, "msg": "无权删除此成员"})
    db.execute("DELETE FROM phone_change_requests WHERE member_id=?", (mid,))
    db.execute("DELETE FROM member_change_requests WHERE member_id=?", (mid,))
    db.execute("DELETE FROM payments WHERE member_id=?", (mid,))
    db.execute("DELETE FROM members WHERE id=?", (mid,))
    log_action(g.admin_id, "删除成员（支部）", f"成员ID：{mid}，姓名：{member['name']}", db=db)
    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "已删除"})


@branch_bp.route("/branch/members/<int:mid>/unbind", methods=["POST"])
@require_admin
def unbind_member(mid):
    db = get_db()
    member = db.execute(
        "SELECT * FROM members WHERE id=? AND branch_id=?", (mid, g.branch_id)
    ).fetchone()
    if not member:
        db.close()
        return jsonify({"code": 403, "msg": "无权操作此成员"})
    db.execute("UPDATE members SET openid=NULL WHERE id=?", (mid,))
    log_action(g.admin_id, "解绑微信（支部）", f"成员ID：{mid}", db=db)
    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "解绑成功"})


# ============================================================
# 本支部缴费看板
# ============================================================

@branch_bp.route("/branch/dashboard", methods=["GET"])
@require_admin
def dashboard():
    # Accept period_ids (comma-separated) or legacy period_id
    period_ids_raw = request.args.get("period_ids", "") or request.args.get("period_id", "")
    period_ids = [int(x) for x in period_ids_raw.split(",") if x.strip().isdigit()]

    db = get_db()
    periods = db.execute("SELECT * FROM periods ORDER BY id DESC").fetchall()

    if not period_ids:
        if not periods:
            db.close()
            return jsonify({
                "code": 0,
                "data": {
                    "period_ids": [], "total": 0, "paid": 0,
                    "unpaid": 0, "amount": 0.0, "details": [], "periods": [],
                },
            })
        period_ids = [periods[0]["id"]]  # default to latest

    placeholders = ",".join("?" * len(period_ids))
    rows = db.execute(
        """SELECT p.id AS payment_id, m.id AS member_id, m.name, m.member_no,
                  m.phone, m.identity, m.branch_id, b.name AS branch_name,
                  p.amount, p.paid, p.pay_type, p.paid_at,
                  p.period_id, pe.name AS period_name
           FROM payments p
           JOIN members m ON p.member_id = m.id
           JOIN branches b ON m.branch_id = b.id
           JOIN periods pe ON p.period_id = pe.id
           WHERE p.period_id IN ({}) AND m.branch_id=?""".format(placeholders),
        (*period_ids, g.branch_id),
    ).fetchall()

    details = [dict(r) for r in rows]
    total = len(details)
    paid_count = sum(1 for d in details if d["paid"])
    amount_sum = sum(d["amount"] for d in details if d["paid"])

    # Branch stats for chart (single branch for branch admins)
    branch_stats = []
    if total > 0:
        rate = round(paid_count / total * 100) if total > 0 else 0
        branch_name_val = details[0]["branch_name"] if details else ""
        branch_stats = [{
            "id": str(g.branch_id), "name": branch_name_val,
            "total": total, "paid": paid_count,
            "unpaid": total - paid_count, "rate": rate,
        }]

    branch = db.execute("SELECT name FROM branches WHERE id=?", (g.branch_id,)).fetchone()
    branch_name = branch["name"] if branch else ""
    db.close()

    return jsonify({
        "code": 0,
        "data": {
            "period_ids": period_ids,
            "branch_name": branch_name,
            "total": total,
            "paid": paid_count,
            "unpaid": total - paid_count,
            "amount": round(amount_sum, 2),
            "details": details,
            "branch_stats": branch_stats,
            "periods": [dict(p) for p in periods],
        },
    })


@branch_bp.route("/branch/confirm/<int:payment_id>", methods=["POST"])
@require_admin
def confirm_payment(payment_id):
    data = request.get_json() or {}
    paid = data.get("paid", 1)

    db = get_db()
    payment = db.execute(
        """SELECT p.* FROM payments p
           JOIN members m ON p.member_id = m.id
           WHERE p.id=? AND m.branch_id=?""",
        (payment_id, g.branch_id),
    ).fetchone()
    if not payment:
        db.close()
        return jsonify({"code": 403, "msg": "无权操作此记录"})

    if paid:
        db.execute(
            "UPDATE payments SET paid=1, pay_type='manual', paid_at=datetime('now','localtime') WHERE id=?",
            (payment_id,),
        )
        log_action(g.admin_id, "手动确认缴费", f"缴费记录ID：{payment_id}", db=db)
    else:
        db.execute(
            "UPDATE payments SET paid=0, pay_type=NULL, paid_at=NULL WHERE id=?",
            (payment_id,),
        )
        log_action(g.admin_id, "撤销缴费确认", f"缴费记录ID：{payment_id}", db=db)

    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "操作成功"})


@branch_bp.route("/branch/confirm_all/<int:period_id>", methods=["POST"])
@require_admin
def confirm_all(period_id):
    db = get_db()
    db.execute(
        """UPDATE payments SET paid=1, pay_type='manual', paid_at=datetime('now','localtime')
           WHERE period_id=? AND paid=0
           AND member_id IN (SELECT id FROM members WHERE branch_id=?)""",
        (period_id, g.branch_id),
    )
    log_action(g.admin_id, "一键确认本支部", f"期数ID：{period_id}", db=db)
    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "已全部确认到账"})


# ============================================================
# 导出 Excel
# ============================================================

# ============================================================
# 换绑手机号审核（super 可见全部，branch 仅本支部）
# ============================================================

@branch_bp.route("/admin/phone_requests", methods=["GET"])
@require_admin
def list_phone_requests():
    status = request.args.get("status", "pending")
    db = get_db()

    sql = """
        SELECT r.id, r.member_id, r.old_phone, r.new_phone, r.status,
               r.reject_reason, r.created_at, r.reviewed_at,
               m.name AS member_name, m.member_no, m.branch_id,
               b.name AS branch_name
        FROM phone_change_requests r
        JOIN members m ON r.member_id = m.id
        LEFT JOIN branches b ON m.branch_id = b.id
        WHERE 1=1
    """
    params = []
    if status and status != "all":
        sql += " AND r.status=?"
        params.append(status)
    if g.role == "branch":
        sql += " AND m.branch_id=?"
        params.append(g.branch_id)
    sql += " ORDER BY r.id DESC LIMIT 200"

    rows = db.execute(sql, params).fetchall()
    db.close()
    return jsonify({"code": 0, "data": [dict(r) for r in rows]})


def _load_request_for_review(db, req_id):
    """取出 pending 申请；分支管理员只能处理本支部的。返回 (row, error_json_or_None)。"""
    row = db.execute(
        """SELECT r.*, m.branch_id
           FROM phone_change_requests r
           JOIN members m ON r.member_id = m.id
           WHERE r.id=?""",
        (req_id,),
    ).fetchone()
    if not row:
        return None, jsonify({"code": 404, "msg": "申请不存在"})
    if row["status"] != "pending":
        return None, jsonify({"code": 400, "msg": "该申请已被处理，无法再次操作"})
    if g.role == "branch" and row["branch_id"] != g.branch_id:
        return None, jsonify({"code": 403, "msg": "无权处理其他支部的申请"})
    return row, None


@branch_bp.route("/admin/phone_requests/<int:req_id>/approve", methods=["POST"])
@require_admin
def approve_phone_request(req_id):
    db = get_db()
    row, err = _load_request_for_review(db, req_id)
    if err:
        db.close()
        return err

    # 审核通过时再检查一次冲突，防并发
    new_phone = row["new_phone"]
    conflict = db.execute(
        "SELECT id FROM members WHERE phone=? AND id!=?",
        (new_phone, row["member_id"]),
    ).fetchone()
    if conflict:
        db.execute(
            """UPDATE phone_change_requests
               SET status='rejected', reject_reason='该手机号已被其他成员占用',
                   reviewed_by=?, reviewed_at=datetime('now','localtime')
               WHERE id=?""",
            (g.admin_id, req_id),
        )
        db.commit()
        db.close()
        return jsonify({"code": 400, "msg": "新手机号已被占用，申请已自动驳回"})

    try:
        db.execute(
            "UPDATE members SET phone=? WHERE id=?",
            (new_phone, row["member_id"]),
        )
        db.execute(
            """UPDATE phone_change_requests
               SET status='approved', reviewed_by=?, reviewed_at=datetime('now','localtime')
               WHERE id=?""",
            (g.admin_id, req_id),
        )
        log_action(g.admin_id, "批准换绑手机", f"申请ID：{req_id}，成员ID：{row['member_id']}", db=db)
        db.commit()
    except Exception as e:
        db.close()
        return jsonify({"code": 500, "msg": f"审核失败：{e}"})
    db.close()
    return jsonify({"code": 0, "msg": "已批准"})


@branch_bp.route("/admin/phone_requests/<int:req_id>/reject", methods=["POST"])
@require_admin
def reject_phone_request(req_id):
    reason = (request.get_json() or {}).get("reason", "").strip() or "管理员未说明原因"
    db = get_db()
    row, err = _load_request_for_review(db, req_id)
    if err:
        db.close()
        return err

    db.execute(
        """UPDATE phone_change_requests
           SET status='rejected', reject_reason=?,
               reviewed_by=?, reviewed_at=datetime('now','localtime')
           WHERE id=?""",
        (reason, g.admin_id, req_id),
    )
    log_action(g.admin_id, "驳回换绑手机", f"申请ID：{req_id}，原因：{reason}", db=db)
    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "已驳回"})


@branch_bp.route("/admin/info_requests", methods=["GET"])
@require_admin
def list_info_requests():
    """列出成员信息修改申请（super 可见全部，branch 仅本支部）。"""
    status = request.args.get("status", "pending")
    db = get_db()

    sql = """
        SELECT r.id, r.member_id, r.changes, r.note, r.status,
               r.reject_reason, r.created_at, r.reviewed_at,
               m.name AS member_name, m.member_no, m.branch_id,
               b.name AS branch_name
        FROM member_change_requests r
        JOIN members m ON r.member_id = m.id
        LEFT JOIN branches b ON m.branch_id = b.id
        WHERE 1=1
    """
    params = []
    if status and status != "all":
        sql += " AND r.status=?"
        params.append(status)
    if g.role == "branch":
        sql += " AND m.branch_id=?"
        params.append(g.branch_id)
    sql += " ORDER BY r.id DESC LIMIT 200"

    rows = db.execute(sql, params).fetchall()
    result = []
    for row in rows:
        r = dict(row)
        try:
            r["changes"] = _json.loads(r["changes"])
        except Exception:
            r["changes"] = {}
        result.append(r)
    db.close()
    return jsonify({"code": 0, "data": result})


def _load_info_request_for_review(db, req_id):
    """取出 pending 申请；分支管理员只能处理本支部的。"""
    row = db.execute(
        """SELECT r.*, m.branch_id AS current_branch_id
           FROM member_change_requests r
           JOIN members m ON r.member_id = m.id
           WHERE r.id=?""",
        (req_id,),
    ).fetchone()
    if not row:
        return None, jsonify({"code": 404, "msg": "申请不存在"})
    if row["status"] != "pending":
        return None, jsonify({"code": 400, "msg": "该申请已被处理，无法再次操作"})
    if g.role == "branch" and row["current_branch_id"] != g.branch_id:
        return None, jsonify({"code": 403, "msg": "无权处理其他支部的申请"})
    return row, None


@branch_bp.route("/admin/info_requests/<int:req_id>/approve", methods=["POST"])
@require_admin
def approve_info_request(req_id):
    db = get_db()
    row, err = _load_info_request_for_review(db, req_id)
    if err:
        db.close()
        return err

    changes = _json.loads(row["changes"])
    update_fields, update_values = [], []

    # 手机号冲突再次检测（防并发）
    if "phone" in changes:
        new_phone = changes["phone"]["new"]
        conflict = db.execute(
            "SELECT id FROM members WHERE phone=? AND id!=?", (new_phone, row["member_id"])
        ).fetchone()
        if conflict:
            db.execute(
                """UPDATE member_change_requests
                   SET status='rejected', reject_reason='手机号已被其他成员占用，申请自动驳回',
                       reviewed_by=?, reviewed_at=datetime('now','localtime')
                   WHERE id=?""",
                (g.admin_id, req_id),
            )
            db.commit()
            db.close()
            return jsonify({"code": 400, "msg": "新手机号已被占用，申请已自动驳回"})
        update_fields.append("phone=?")
        update_values.append(new_phone)

    if "identity" in changes:
        update_fields.append("identity=?")
        update_values.append(changes["identity"]["new"])

    if "branch_id" in changes:
        if g.role != "super":
            db.close()
            return jsonify({"code": 403, "msg": "支部管理员无权审批跨支部转移申请"})
        update_fields.append("branch_id=?")
        update_values.append(changes["branch_id"]["new"])

    if "amount" in changes:
        update_fields.append("amount=?")
        update_values.append(changes["amount"]["new"])

    try:
        if update_fields:
            update_values.append(row["member_id"])
            db.execute(
                f"UPDATE members SET {','.join(update_fields)} WHERE id=?", update_values
            )
        db.execute(
            """UPDATE member_change_requests
               SET status='approved', reviewed_by=?, reviewed_at=datetime('now','localtime')
               WHERE id=?""",
            (g.admin_id, req_id),
        )
        log_action(g.admin_id, "批准信息修改申请", f"申请ID：{req_id}，成员ID：{row['member_id']}", db=db)
        db.commit()
    except Exception as e:
        db.close()
        return jsonify({"code": 500, "msg": f"审核失败：{e}"})
    db.close()
    return jsonify({"code": 0, "msg": "已批准"})


@branch_bp.route("/admin/info_requests/<int:req_id>/reject", methods=["POST"])
@require_admin
def reject_info_request(req_id):
    reason = (request.get_json() or {}).get("reason", "").strip() or "管理员未说明原因"
    db = get_db()
    row, err = _load_info_request_for_review(db, req_id)
    if err:
        db.close()
        return err

    db.execute(
        """UPDATE member_change_requests
           SET status='rejected', reject_reason=?,
               reviewed_by=?, reviewed_at=datetime('now','localtime')
           WHERE id=?""",
        (reason, g.admin_id, req_id),
    )
    log_action(g.admin_id, "驳回信息修改申请", f"申请ID：{req_id}，原因：{reason}", db=db)
    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "已驳回"})


@branch_bp.route("/admin/info_requests/count", methods=["GET"])
@require_admin
def count_pending_info_requests():
    db = get_db()
    sql = """SELECT COUNT(*) FROM member_change_requests r
             JOIN members m ON r.member_id = m.id
             WHERE r.status='pending'"""
    params = []
    if g.role == "branch":
        sql += " AND m.branch_id=?"
        params.append(g.branch_id)
    n = db.execute(sql, params).fetchone()[0]
    db.close()
    return jsonify({"code": 0, "data": {"pending": n}})


@branch_bp.route("/admin/phone_requests/count", methods=["GET"])
@require_admin
def count_pending_phone_requests():
    db = get_db()
    sql = """SELECT COUNT(*) FROM phone_change_requests r
             JOIN members m ON r.member_id = m.id
             WHERE r.status='pending'"""
    params = []
    if g.role == "branch":
        sql += " AND m.branch_id=?"
        params.append(g.branch_id)
    n = db.execute(sql, params).fetchone()[0]
    db.close()
    return jsonify({"code": 0, "data": {"pending": n}})


@branch_bp.route("/branch/export", methods=["GET"])
@require_admin
def export_excel():
    # 支持多期数筛选（逗号分隔）
    period_ids_raw = request.args.get("period_ids", "") or request.args.get("period_id", "")
    period_ids = [int(x) for x in period_ids_raw.split(",") if x.strip().isdigit()]

    if not period_ids:
        return jsonify({"code": 400, "msg": "请选择期数"})

    db = get_db()
    period_rows = db.execute(
        "SELECT name FROM periods WHERE id IN ({}) ORDER BY id".format(",".join("?" * len(period_ids))),
        period_ids,
    ).fetchall()
    if not period_rows:
        db.close()
        return jsonify({"code": 404, "msg": "期数不存在"})

    rows = db.execute(
        """SELECT pe.name AS period_name, m.name, m.member_no, m.phone, m.identity, m.person_type,
                  p.amount, p.paid, p.paid_at, p.pay_type
           FROM payments p
           JOIN members m ON p.member_id = m.id
           JOIN periods pe ON p.period_id = pe.id
           WHERE p.period_id IN ({}) AND m.branch_id=?
           ORDER BY p.period_id, m.member_no""".format(",".join("?" * len(period_ids))),
        (*period_ids, g.branch_id),
    ).fetchall()
    db.close()

    period_label = "+".join(r["name"] for r in period_rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = period_label[:31]

    headers = ["期数", "姓名", "工号/学号", "手机号", "党员身份", "身份(教师/学生)", "应缴金额(元)", "状态", "缴费时间", "缴费方式"]
    ws.append(headers)

    red_fill = PatternFill(fill_type="solid", fgColor="C0292B")
    white_font = Font(color="FFFFFF", bold=True)
    paid_fill = PatternFill(fill_type="solid", fgColor="D5F5E3")
    unpaid_fill = PatternFill(fill_type="solid", fgColor="FADBD8")
    for cell in ws[1]:
        cell.fill = red_fill
        cell.font = white_font

    pay_type_map = {"manual": "手动确认", "wxpay": "微信支付", "mock": "测试支付"}
    for row in rows:
        ws.append([
            row["period_name"],
            row["name"], row["member_no"], row["phone"], row["identity"],
            row["person_type"] or "",
            row["amount"],
            "已缴" if row["paid"] else "未缴",
            row["paid_at"] or "",
            pay_type_map.get(row["pay_type"], "") if row["pay_type"] else "",
        ])
        for cell in ws[ws.max_row]:
            cell.fill = paid_fill if row["paid"] else unpaid_fill

    for i, w in enumerate([14, 10, 15, 14, 12, 14, 14, 8, 20, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"{period_label}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

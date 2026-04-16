import io
import re
import bcrypt
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from flask import Blueprint, request, jsonify, g, send_file
from database import get_db
from auth import require_super, require_admin, log_action

super_bp = Blueprint("super", __name__)

_PHONE_RE = re.compile(r"^\d{11}$")


def _validate_member_payload(data, require_branch=True):
    """校验成员表单：返回错误消息或 None。"""
    for field in ["name", "member_no", "phone", "identity"]:
        v = data.get(field)
        if v is None or str(v).strip() == "":
            return f"字段 {field} 不能为空"
    if require_branch and not data.get("branch_id"):
        return "字段 branch_id 不能为空"
    if not _PHONE_RE.match(str(data.get("phone", "")).strip()):
        return "手机号必须为11位数字"
    if data.get("amount") is None or str(data.get("amount")).strip() == "":
        return "应缴金额不能为空"
    try:
        amt = float(data["amount"])
    except (TypeError, ValueError):
        return "应缴金额格式错误"
    if amt <= 0:
        return "应缴金额须为正数"
    if amt > 100000:
        return "应缴金额异常（>100000）"
    # 统一归一到 2 位小数，避免浮点脏数据入库
    data["amount"] = round(amt, 2)
    status = data.get("status", "active")
    if status not in ("active", "suspended", "transferred"):
        return "状态值非法"
    return None


def _friendly_unique_msg(err_text):
    """将 SQLite UNIQUE 冲突翻译成友好提示。"""
    s = str(err_text)
    if "UNIQUE" in s and "branches.name" in s:
        return "支部名称已存在"
    if "UNIQUE" in s and "members.member_no" in s:
        return "工号/学号已存在"
    if "UNIQUE" in s and "members.phone" in s:
        return "手机号已被占用"
    if "UNIQUE" in s and "admins.username" in s:
        return "用户名已存在"
    return None


# ============================================================
# 支部管理
# ============================================================

@super_bp.route("/super/branches", methods=["GET"])
@require_super
def list_branches():
    db = get_db()
    rows = db.execute("SELECT * FROM branches ORDER BY id").fetchall()
    db.close()
    return jsonify({"code": 0, "data": [dict(r) for r in rows]})


@super_bp.route("/super/branches", methods=["POST"])
@require_super
def create_branch():
    data = request.get_json() or {}
    name = data.get("name", "").strip()
    leader = data.get("leader", "").strip()
    if not name:
        return jsonify({"code": 400, "msg": "支部名称不能为空"})
    db = get_db()
    if db.execute("SELECT id FROM branches WHERE name=?", (name,)).fetchone():
        db.close()
        return jsonify({"code": 400, "msg": f"支部「{name}」已存在"})
    try:
        db.execute("INSERT INTO branches (name, leader) VALUES (?, ?)", (name, leader))
        log_action(g.admin_id, "创建支部", f"支部名称：{name}", db=db)
        db.commit()
    except Exception as e:
        db.close()
        return jsonify({"code": 500, "msg": str(e)})
    db.close()
    return jsonify({"code": 0, "msg": "创建成功"})


@super_bp.route("/super/branches/<int:bid>", methods=["PUT"])
@require_super
def update_branch(bid):
    data = request.get_json() or {}
    name = data.get("name", "").strip()
    leader = data.get("leader", "").strip()
    if not name:
        return jsonify({"code": 400, "msg": "支部名称不能为空"})
    db = get_db()
    conflict = db.execute(
        "SELECT id FROM branches WHERE name=? AND id!=?", (name, bid)
    ).fetchone()
    if conflict:
        db.close()
        return jsonify({"code": 400, "msg": f"支部「{name}」已存在"})
    db.execute("UPDATE branches SET name=?, leader=? WHERE id=?", (name, leader, bid))
    log_action(g.admin_id, "编辑支部", f"支部ID：{bid}", db=db)
    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "更新成功"})


@super_bp.route("/super/branches/<int:bid>", methods=["DELETE"])
@require_super
def delete_branch(bid):
    db = get_db()
    # 检查支部下是否仍有成员
    member_count = db.execute(
        "SELECT COUNT(*) FROM members WHERE branch_id=?", (bid,)
    ).fetchone()[0]
    if member_count > 0:
        db.close()
        return jsonify({"code": 400, "msg": f"该支部下有 {member_count} 名成员，请先转移或删除成员后再删除支部"})
    try:
        db.execute("DELETE FROM branches WHERE id=?", (bid,))
        log_action(g.admin_id, "删除支部", f"支部ID：{bid}", db=db)
        db.commit()
    except Exception as e:
        db.close()
        return jsonify({"code": 500, "msg": str(e)})
    db.close()
    return jsonify({"code": 0, "msg": "删除成功"})


# ============================================================
# 管理员账号管理
# ============================================================

@super_bp.route("/super/admins", methods=["GET"])
@require_super
def list_admins():
    db = get_db()
    rows = db.execute("""
        SELECT a.id, a.username, a.role, a.branch_id, a.is_active, a.created_at,
               b.name AS branch_name
        FROM admins a
        LEFT JOIN branches b ON a.branch_id = b.id
        ORDER BY a.id
    """).fetchall()
    db.close()
    return jsonify({"code": 0, "data": [dict(r) for r in rows]})


@super_bp.route("/super/admins", methods=["POST"])
@require_super
def create_admin():
    data = request.get_json() or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    role = data.get("role", "branch")
    branch_id = data.get("branch_id") or None

    if not username or not password:
        return jsonify({"code": 400, "msg": "用户名和密码不能为空"})
    if len(password) < 8:
        return jsonify({"code": 400, "msg": "密码长度不能少于8位"})
    if role == "branch" and not branch_id:
        return jsonify({"code": 400, "msg": "普通管理员必须绑定支部"})

    hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    db = get_db()
    try:
        db.execute(
            "INSERT INTO admins (username, password, role, branch_id) VALUES (?, ?, ?, ?)",
            (username, hashed, role, branch_id),
        )
        log_action(g.admin_id, "创建管理员", f"用户名：{username}，角色：{role}", db=db)
        db.commit()
    except Exception as e:
        db.close()
        return jsonify({"code": 400, "msg": _friendly_unique_msg(e) or "用户名已存在"})
    db.close()
    return jsonify({"code": 0, "msg": "创建成功"})


@super_bp.route("/super/admins/<int:aid>", methods=["PUT"])
@require_super
def update_admin(aid):
    data = request.get_json() or {}
    db = get_db()
    fields, values = [], []

    if "username" in data:
        username = (data.get("username") or "").strip()
        if not username:
            db.close()
            return jsonify({"code": 400, "msg": "用户名不能为空"})
        fields.append("username=?")
        values.append(username)
    if data.get("password"):
        if len(data["password"]) < 8:
            db.close()
            return jsonify({"code": 400, "msg": "密码长度不能少于8位"})
        hashed = bcrypt.hashpw(data["password"].encode(), bcrypt.gensalt()).decode()
        fields.append("password=?")
        values.append(hashed)
    if "branch_id" in data:
        fields.append("branch_id=?")
        values.append(data["branch_id"] or None)
    if "is_active" in data:
        fields.append("is_active=?")
        values.append(1 if data["is_active"] else 0)

    if fields:
        values.append(aid)
        try:
            db.execute(f"UPDATE admins SET {','.join(fields)} WHERE id=?", values)
            log_action(g.admin_id, "编辑管理员", f"管理员ID：{aid}", db=db)
            db.commit()
        except Exception as e:
            db.close()
            return jsonify({"code": 400, "msg": _friendly_unique_msg(e) or "保存失败"})

    db.close()
    return jsonify({"code": 0, "msg": "更新成功"})


@super_bp.route("/super/admins/<int:aid>", methods=["DELETE"])
@require_super
def delete_admin(aid):
    if aid == g.admin_id:
        return jsonify({"code": 400, "msg": "不能删除自己"})
    db = get_db()
    # 防止删除最后一个超级管理员
    target = db.execute("SELECT role FROM admins WHERE id=?", (aid,)).fetchone()
    if target and target["role"] == "super":
        super_count = db.execute(
            "SELECT COUNT(*) FROM admins WHERE role='super'"
        ).fetchone()[0]
        if super_count <= 1:
            db.close()
            return jsonify({"code": 400, "msg": "不能删除最后一个超级管理员"})
    db.execute("DELETE FROM admins WHERE id=?", (aid,))
    log_action(g.admin_id, "删除管理员", f"管理员ID：{aid}", db=db)
    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "删除成功"})


@super_bp.route("/super/admins/<int:aid>/logs", methods=["GET"])
@require_super
def admin_logs(aid):
    db = get_db()
    rows = db.execute(
        "SELECT * FROM admin_logs WHERE admin_id=? ORDER BY created_at DESC LIMIT 100",
        (aid,),
    ).fetchall()
    db.close()
    return jsonify({"code": 0, "data": [dict(r) for r in rows]})


# ============================================================
# 成员管理（全院）
# ============================================================

@super_bp.route("/super/members", methods=["GET"])
@require_super
def list_members():
    branch_id = request.args.get("branch_id")
    identity = request.args.get("identity")
    status = request.args.get("status")

    sql = """
        SELECT m.*, b.name AS branch_name
        FROM members m
        LEFT JOIN branches b ON m.branch_id = b.id
        WHERE 1=1
    """
    params = []
    if branch_id:
        sql += " AND m.branch_id=?"
        params.append(branch_id)
    if identity:
        sql += " AND m.identity=?"
        params.append(identity)
    if status:
        sql += " AND m.status=?"
        params.append(status)
    sql += " ORDER BY m.id"

    db = get_db()
    rows = db.execute(sql, params).fetchall()
    db.close()
    return jsonify({"code": 0, "data": [dict(r) for r in rows]})


@super_bp.route("/super/members", methods=["POST"])
@require_super
def create_member():
    data = request.get_json() or {}
    err = _validate_member_payload(data, require_branch=True)
    if err:
        return jsonify({"code": 400, "msg": err})

    db = get_db()
    try:
        db.execute(
            """INSERT INTO members (name, member_no, phone, identity, person_type, branch_id, amount, status, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                data["name"].strip(), str(data["member_no"]).strip(),
                str(data["phone"]).strip(), data["identity"].strip(),
                (data.get("person_type") or "").strip(),
                data["branch_id"], float(data["amount"]),
                data.get("status", "active"), (data.get("notes") or "").strip(),
            ),
        )
        log_action(g.admin_id, "新增成员", f"姓名：{data['name']}", db=db)
        db.commit()
    except Exception as e:
        db.close()
        return jsonify({"code": 400, "msg": _friendly_unique_msg(e) or "工号/学号或手机号已存在"})
    db.close()
    return jsonify({"code": 0, "msg": "新增成功"})


@super_bp.route("/super/members/<int:mid>", methods=["PUT"])
@require_super
def update_member(mid):
    data = request.get_json() or {}
    err = _validate_member_payload(data, require_branch=True)
    if err:
        return jsonify({"code": 400, "msg": err})

    db = get_db()
    try:
        db.execute(
            """UPDATE members SET name=?, member_no=?, phone=?, identity=?, person_type=?,
               branch_id=?, amount=?, status=?, notes=? WHERE id=?""",
            (
                data["name"].strip(), str(data["member_no"]).strip(),
                str(data["phone"]).strip(), data["identity"].strip(),
                (data.get("person_type") or "").strip(),
                data["branch_id"], float(data["amount"]),
                data.get("status", "active"), (data.get("notes") or "").strip(), mid,
            ),
        )
        log_action(g.admin_id, "编辑成员", f"成员ID：{mid}", db=db)
        db.commit()
    except Exception as e:
        db.close()
        return jsonify({"code": 400, "msg": _friendly_unique_msg(e) or "保存失败：工号/学号或手机号冲突"})
    db.close()
    return jsonify({"code": 0, "msg": "更新成功"})


@super_bp.route("/super/members/<int:mid>", methods=["DELETE"])
@require_super
def delete_member(mid):
    db = get_db()
    db.execute("DELETE FROM phone_change_requests WHERE member_id=?", (mid,))
    db.execute("DELETE FROM member_change_requests WHERE member_id=?", (mid,))
    db.execute("DELETE FROM payments WHERE member_id=?", (mid,))
    db.execute("DELETE FROM members WHERE id=?", (mid,))
    log_action(g.admin_id, "删除成员", f"成员ID：{mid}", db=db)
    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "删除成功"})


@super_bp.route("/super/members/<int:mid>/unbind", methods=["POST"])
@require_super
def unbind_member(mid):
    db = get_db()
    db.execute("UPDATE members SET openid=NULL WHERE id=?", (mid,))
    log_action(g.admin_id, "解绑成员微信", f"成员ID：{mid}", db=db)
    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "解绑成功"})


@super_bp.route("/super/members/template", methods=["GET"])
@require_admin
def member_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成员导入模板"
    headers = ["姓名", "工号/学号", "手机号", "党员身份", "身份(教师/学生)", "支部名称", "应缴金额(元)", "状态(正常/停缴/已转出)", "备注"]
    ws.append(headers)
    ws.append(["张三", "20240001", "13800138000", "正式党员", "教师", "计算机支部", "5", "正常", ""])

    red_fill = PatternFill(fill_type="solid", fgColor="C0292B")
    white_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = red_fill
        cell.font = white_font

    for i, w in enumerate([10, 15, 14, 12, 14, 15, 14, 18, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="成员导入模板.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@super_bp.route("/super/members/import", methods=["POST"])
@require_super
def import_members():
    if "file" not in request.files:
        return jsonify({"code": 400, "msg": "未找到文件"})

    file = request.files["file"]
    file.seek(0, 2)
    if file.tell() > 5 * 1024 * 1024:
        return jsonify({"code": 400, "msg": "文件大小不能超过 5MB"})
    file.seek(0)
    try:
        wb = openpyxl.load_workbook(file)
    except Exception:
        return jsonify({"code": 400, "msg": "文件格式错误，请上传 .xlsx 文件"})

    ws = wb.active
    if ws.max_row - 1 > 1000:
        return jsonify({"code": 400, "msg": "单次导入不能超过 1000 条"})
    db = get_db()
    success_count, skip_count, error_rows = 0, 0, []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        try:
            name        = str(row[0]).strip()
            member_no   = str(row[1]).strip()
            phone       = str(row[2]).strip()
            identity    = str(row[3]).strip()
            person_type = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            branch_name = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            amount      = float(row[6] or 0) if len(row) > 6 and row[6] is not None else 0
            status_raw  = str(row[7]).strip() if len(row) > 7 and row[7] else "正常"
            notes       = str(row[8]).strip() if len(row) > 8 and row[8] else ""
            _STATUS_MAP = {"正常": "active", "停缴": "suspended", "已转出": "transferred",
                           "active": "active", "suspended": "suspended", "transferred": "transferred"}
            status = _STATUS_MAP.get(status_raw)

            if not name or not member_no or not phone:
                error_rows.append(f"第{i}行：姓名、工号/学号、手机号不能为空")
                continue
            if not _PHONE_RE.match(phone):
                error_rows.append(f"第{i}行：手机号「{phone}」必须为11位数字")
                continue
            if amount <= 0:
                error_rows.append(f"第{i}行：应缴金额须为正数")
                continue
            if amount > 100000:
                error_rows.append(f"第{i}行：应缴金额异常（>100000）")
                continue
            if status is None:
                error_rows.append(f"第{i}行：状态值「{status_raw}」非法，请填写正常/停缴/已转出")
                continue
            amount = round(amount, 2)

            branch = db.execute("SELECT id FROM branches WHERE name=?", (branch_name,)).fetchone()
            if not branch:
                error_rows.append(f"第{i}行：支部「{branch_name}」不存在")
                continue

            cursor = db.execute(
                """INSERT OR IGNORE INTO members
                   (name, member_no, phone, identity, person_type, branch_id, amount, status, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (name, member_no, phone, identity, person_type, branch["id"], amount, status, notes),
            )
            if cursor.rowcount > 0:
                success_count += 1
            else:
                skip_count += 1  # 工号或手机号已存在，跳过
        except Exception as e:
            error_rows.append(f"第{i}行：{str(e)}")

    log_action(g.admin_id, "批量导入成员", f"成功{success_count}条，跳过{skip_count}条", db=db)
    db.commit()
    db.close()

    parts = [f"成功导入 {success_count} 条"]
    if skip_count > 0:
        parts.append(f"{skip_count} 条已存在跳过")
    return jsonify({
        "code": 0,
        "msg": "，".join(parts),
        "errors": error_rows,
    })


# ============================================================
# 期数管理
# ============================================================

@super_bp.route("/super/periods", methods=["GET"])
@require_super
def list_periods():
    db = get_db()
    rows = db.execute("SELECT * FROM periods ORDER BY id DESC").fetchall()
    db.close()
    return jsonify({"code": 0, "data": [dict(r) for r in rows]})


@super_bp.route("/super/periods", methods=["POST"])
@require_super
def create_period():
    data = request.get_json() or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"code": 400, "msg": "期数名称不能为空"})

    db = get_db()
    # 检查期数名称是否已存在
    existing = db.execute("SELECT id FROM periods WHERE name=?", (name,)).fetchone()
    if existing:
        db.close()
        return jsonify({"code": 400, "msg": f"期数「{name}」已存在"})

    try:
        cursor = db.execute("INSERT INTO periods (name) VALUES (?)", (name,))
        period_id = cursor.lastrowid

        members = db.execute(
            "SELECT id, amount FROM members WHERE status='active'"
        ).fetchall()
        for m in members:
            db.execute(
                "INSERT OR IGNORE INTO payments (member_id, period_id, amount) VALUES (?, ?, ?)",
                (m["id"], period_id, m["amount"]),
            )

        log_action(g.admin_id, "新建期数", f"期数：{name}，生成{len(members)}条记录", db=db)
        db.commit()
    except Exception as e:
        db.close()
        return jsonify({"code": 500, "msg": str(e)})
    db.close()
    return jsonify({"code": 0, "msg": f"已新建期数，为 {len(members)} 名成员生成缴费记录"})


@super_bp.route("/super/periods/<int:pid>", methods=["DELETE"])
@require_super
def delete_period(pid):
    db = get_db()
    db.execute("DELETE FROM payments WHERE period_id=?", (pid,))
    db.execute("DELETE FROM periods WHERE id=?", (pid,))
    log_action(g.admin_id, "删除期数", f"期数ID：{pid}", db=db)
    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "删除成功"})


# ============================================================
# 缴费看板（全院）
# ============================================================

@super_bp.route("/super/dashboard", methods=["GET"])
@require_super
def dashboard():
    # Accept period_ids (comma-separated) or legacy period_id
    period_ids_raw = request.args.get("period_ids", "") or request.args.get("period_id", "")
    period_ids = [int(x) for x in period_ids_raw.split(",") if x.strip().isdigit()]
    branch_id = request.args.get("branch_id")

    db = get_db()
    periods = db.execute("SELECT * FROM periods ORDER BY id DESC").fetchall()
    branches = db.execute("SELECT * FROM branches ORDER BY id").fetchall()

    if not period_ids:
        if not periods:
            db.close()
            return jsonify({
                "code": 0,
                "data": {
                    "period_ids": [], "total": 0, "paid": 0,
                    "unpaid": 0, "amount": 0.0, "details": [],
                    "periods": [], "branches": [],
                },
            })
        period_ids = [periods[0]["id"]]  # default to latest

    placeholders = ",".join("?" * len(period_ids))
    sql = """
        SELECT p.id AS payment_id, m.id AS member_id, m.name, m.member_no,
               m.phone, m.identity, b.id AS branch_id, b.name AS branch_name,
               p.amount, p.paid, p.pay_type, p.paid_at,
               p.period_id, pe.name AS period_name
        FROM payments p
        JOIN members m ON p.member_id = m.id
        JOIN branches b ON m.branch_id = b.id
        JOIN periods pe ON p.period_id = pe.id
        WHERE p.period_id IN ({})
    """.format(placeholders)
    params = list(period_ids)
    if branch_id:
        sql += " AND m.branch_id=?"
        params.append(branch_id)

    rows = db.execute(sql, params).fetchall()
    details = [dict(r) for r in rows]
    total = len(details)
    paid_count = sum(1 for d in details if d["paid"])
    amount_sum = sum(d["amount"] for d in details if d["paid"])

    # Per-branch stats for chart
    branch_map = {}
    for d in details:
        key = str(d.get("branch_id") or "unknown")
        if key not in branch_map:
            branch_map[key] = {"id": key, "name": d.get("branch_name") or "未知", "total": 0, "paid": 0}
        branch_map[key]["total"] += 1
        if d["paid"]:
            branch_map[key]["paid"] += 1
    branch_stats = []
    for b in branch_map.values():
        b["unpaid"] = b["total"] - b["paid"]
        b["rate"] = round(b["paid"] / b["total"] * 100) if b["total"] > 0 else 0
        branch_stats.append(b)

    db.close()

    return jsonify({
        "code": 0,
        "data": {
            "period_ids": period_ids,
            "total": total,
            "paid": paid_count,
            "unpaid": total - paid_count,
            "amount": round(amount_sum, 2),
            "details": details,
            "branch_stats": branch_stats,
            "periods": [dict(p) for p in periods],
            "branches": [dict(b) for b in branches],
        },
    })


@super_bp.route("/super/confirm/<int:payment_id>", methods=["POST"])
@require_admin
def confirm_payment(payment_id):
    data = request.get_json() or {}
    paid = data.get("paid", 1)

    db = get_db()
    # 超级管理员无限制；普通管理员只能操作本支部
    if g.role == "branch":
        payment = db.execute(
            """SELECT p.* FROM payments p
               JOIN members m ON p.member_id = m.id
               WHERE p.id=? AND m.branch_id=?""",
            (payment_id, g.branch_id),
        ).fetchone()
        if not payment:
            db.close()
            return jsonify({"code": 403, "msg": "无权操作此记录"})

    if paid:
        db.execute(
            "UPDATE payments SET paid=1, pay_type='manual', paid_at=datetime('now','localtime') WHERE id=?",
            (payment_id,),
        )
        log_action(g.admin_id, "手动确认缴费", f"缴费记录ID：{payment_id}", db=db)
    else:
        db.execute(
            "UPDATE payments SET paid=0, pay_type=NULL, paid_at=NULL WHERE id=?",
            (payment_id,),
        )
        log_action(g.admin_id, "撤销缴费确认", f"缴费记录ID：{payment_id}", db=db)

    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "操作成功"})


@super_bp.route("/super/confirm_all/<int:period_id>", methods=["POST"])
@require_super
def confirm_all(period_id):
    db = get_db()
    db.execute(
        "UPDATE payments SET paid=1, pay_type='manual', paid_at=datetime('now','localtime') WHERE period_id=? AND paid=0",
        (period_id,),
    )
    log_action(g.admin_id, "一键全院确认", f"期数ID：{period_id}", db=db)
    db.commit()
    db.close()
    return jsonify({"code": 0, "msg": "已全院确认到账"})


# ============================================================
# 导出 Excel
# ============================================================

@super_bp.route("/super/export", methods=["GET"])
@require_super
def export_excel():
    # 支持多期数、多支部筛选（逗号分隔）
    period_ids_raw = request.args.get("period_ids", "") or request.args.get("period_id", "")
    period_ids = [int(x) for x in period_ids_raw.split(",") if x.strip().isdigit()]
    branch_ids_raw = request.args.get("branch_ids", "") or request.args.get("branch_id", "")
    branch_ids = [int(x) for x in branch_ids_raw.split(",") if x.strip().isdigit()]

    if not period_ids:
        return jsonify({"code": 400, "msg": "请选择期数"})

    db = get_db()
    period_rows = db.execute(
        "SELECT name FROM periods WHERE id IN ({}) ORDER BY id".format(",".join("?" * len(period_ids))),
        period_ids,
    ).fetchall()
    if not period_rows:
        db.close()
        return jsonify({"code": 404, "msg": "期数不存在"})

    sql = """
        SELECT pe.name AS period_name, m.name, m.member_no, m.phone, m.identity, m.person_type,
               b.name AS branch_name, p.amount, p.paid, p.paid_at, p.pay_type
        FROM payments p
        JOIN members m ON p.member_id = m.id
        JOIN branches b ON m.branch_id = b.id
        JOIN periods pe ON p.period_id = pe.id
        WHERE p.period_id IN ({})
    """.format(",".join("?" * len(period_ids)))
    params = list(period_ids)
    if branch_ids:
        sql += " AND m.branch_id IN ({})".format(",".join("?" * len(branch_ids)))
        params.extend(branch_ids)
    sql += " ORDER BY p.period_id, b.id, m.member_no"

    rows = db.execute(sql, params).fetchall()
    db.close()

    period_label = "+".join(r["name"] for r in period_rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = period_label[:31]  # Excel sheet name limit

    headers = ["期数", "姓名", "工号/学号", "手机号", "党员身份", "身份(教师/学生)", "所在支部", "应缴金额(元)", "状态", "缴费时间", "缴费方式"]
    ws.append(headers)

    red_fill = PatternFill(fill_type="solid", fgColor="C0292B")
    white_font = Font(color="FFFFFF", bold=True)
    paid_fill = PatternFill(fill_type="solid", fgColor="D5F5E3")
    unpaid_fill = PatternFill(fill_type="solid", fgColor="FADBD8")
    for cell in ws[1]:
        cell.fill = red_fill
        cell.font = white_font

    pay_type_map = {"manual": "手动确认", "wxpay": "微信支付", "mock": "测试支付"}
    for row in rows:
        ws.append([
            row["period_name"],
            row["name"], row["member_no"], row["phone"], row["identity"],
            row["person_type"] or "",
            row["branch_name"], row["amount"],
            "已缴" if row["paid"] else "未缴",
            row["paid_at"] or "",
            pay_type_map.get(row["pay_type"], "") if row["pay_type"] else "",
        ])
        fill = paid_fill if row["paid"] else unpaid_fill
        for cell in ws[ws.max_row]:
            cell.fill = fill

    col_widths = [14, 10, 15, 14, 12, 14, 20, 14, 8, 20, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"{period_label}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
